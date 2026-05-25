"""
TestCaseExecutor — Excel 用例 → 执行并标注结果

逐页逐步执行 approved 的测试用例，记录 PASS/FAIL 和表单校验错误。
"""
from __future__ import annotations

import os
import tempfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .browser import _BrowserSession
from .browser_test_engine import ActionGenerator, ActionExecutor, PageDiscoverer, TestConfig

PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
SKIP_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
BOLD_FONT = Font(bold=True, size=10)

HEADER_MAP = {
    "case_id": 1, "case_title": 2, "precondition": 3, "case_type": 4,
    "page_url": 5, "step_no": 6, "current_page": 7, "target_page": 8,
    "element_role": 9, "element_text": 10, "element_type": 11, "element_index": 12,
    "action": 13, "input_value": 14, "expected_behavior": 15,
    "status": 16, "result": 17, "error_message": 18, "duration_ms": 19,
}


class TestCaseExecutor:
    """从 Excel 读取 approved 用例 → 执行 → 写回 Excel。"""

    def __init__(self, xlsx_path: str, config: TestConfig):
        self._xlsx_path = xlsx_path
        self._config = config
        self._session: Optional[_BrowserSession] = None
        self._executor: Optional[ActionExecutor] = None
        self._progress_callback: Optional[callable] = None
        self._stop_requested = False
        self._result_path = ""
        self._last_progress: Dict[str, Any] = {
            "running": False, "status": "not_started",
            "done": 0, "total": 0, "passed": 0, "failed": 0,
            "result_path": "", "video_path": "", "error": "",
        }

    def on_progress(self, callback: callable):
        self._progress_callback = callback

    def stop(self):
        self._stop_requested = True

    async def execute(self, auto_approve: bool = False) -> str:
        """Execute cases grouped by case_id, with FP page replay support."""
        self._last_progress["running"] = True
        self._last_progress["status"] = "running"
        self._last_progress["error"] = ""

        wb = load_workbook(self._xlsx_path)
        ws = wb.active

        # Parse approved rows, grouped by case_id
        cases: Dict[str, List[Dict[str, Any]]] = {}
        for row in range(2, ws.max_row + 1):
            status = (ws.cell(row=row, column=HEADER_MAP["status"]).value or "").strip().upper()
            if status != "APPROVED":
                continue
            task = {}
            for key, col in HEADER_MAP.items():
                task[key] = ws.cell(row=row, column=col).value
            task["_row"] = row
            cid = str(task.get("case_id", "") or "default")
            if cid not in cases:
                cases[cid] = []
            cases[cid].append(task)

        if not cases and auto_approve:
            for row in range(2, ws.max_row + 1):
                status = (ws.cell(row=row, column=HEADER_MAP["status"]).value or "").strip().upper()
                if status != "PENDING":
                    continue
                task = {}
                for key, col in HEADER_MAP.items():
                    task[key] = ws.cell(row=row, column=col).value
                task["_row"] = row
                cid = str(task.get("case_id", "") or "default")
                if cid not in cases:
                    cases[cid] = []
                cases[cid].append(task)

        if not cases:
            self._last_progress["error"] = "No APPROVED or PENDING test cases found"
            self._last_progress["running"] = False
            self._last_progress["status"] = "finished"
            raise ValueError(self._last_progress["error"])

        # Start browser + recording
        self._session = _BrowserSession()
        video_dir = os.path.join(
            os.environ.get("AIPLAT_HOME", os.path.expanduser("~/.aiplat")),
            "videos", f"test_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}",
        )
        os.makedirs(video_dir, exist_ok=True)
        await self._session._lazy_start(record_video_dir=video_dir)

        # Login if configured
        if self._config.login_url and self._config.accounts:
            from .browser_test_engine import SessionManager
            sm = SessionManager(self._session, self._config)
            if not await sm.ensure_login():
                raise RuntimeError("Login failed with all accounts")

        total_steps = sum(len(t) for t in cases.values())
        done = passed = failed = 0
        self._last_progress["total"] = total_steps
        self._notify("started", {"total_steps": total_steps})

        import time

        for cid, tasks in cases.items():
            if self._stop_requested:
                break

            # Determine base page: strip #fp_* / #card_* suffix for FP pages
            first_url = str(tasks[0].get("page_url", "") or "")
            base_url = self._strip_fp_suffix(first_url)

            # Navigate to base URL
            try:
                await self._session.goto(base_url)
                await self._session.wait(2000)
            except Exception:
                for t in tasks:
                    self._write_result(ws, t["_row"], "FAIL", "Page load failed", 0)
                    failed += 1; done += 1
                continue

            # Execute all steps in this case in order
            for i, task in enumerate(tasks):
                if self._stop_requested:
                    break

                action_name = str(task.get("action", "") or "")
                role = str(task.get("element_role", "") or "")
                input_val = str(task.get("input_value", "") or "")
                el_idx = int(task.get("element_index", 0) or 0)

                # For FP pages: replay preceding steps to reach current state
                page_url = str(task.get("page_url", "") or "")
                if i > 0 and self._is_fp_url(page_url):
                    prev_url = self._strip_fp_suffix(page_url)
                    if prev_url != base_url:
                        await self._session.goto(prev_url)
                        await self._session.wait(2000)
                        # Replay previous steps (without result recording)
                        for j in range(i):
                            prev = tasks[j]
                            await self._replay_step(prev)
                        base_url = prev_url

                t0 = time.time()
                try:
                    if action_name in ("点击", "click_target"):
                        if el_idx > 0:
                            await self._session.click_index(el_idx)
                        else:
                            await self._session.click_target({"role": role, "text_contains": str(task.get("element_text", "") or "")[:20]})
                        await self._session.wait(500)
                        form_errors = await self._session.detect_form_errors()
                        if form_errors:
                            errs = [e.get("text", "")[:80] for e in form_errors[:3]]
                            self._write_result(ws, task["_row"], "FAIL", " | ".join(errs), (time.time() - t0) * 1000)
                            failed += 1
                        else:
                            self._write_result(ws, task["_row"], "PASS", "", (time.time() - t0) * 1000)
                            passed += 1
                    elif action_name in ("输入", "type_target"):
                        if el_idx > 0:
                            await self._session.type_index(el_idx, input_val or "test")
                        else:
                            await self._session.type_target({"role": role, "placeholder": str(task.get("element_text", "") or "").lower()}, input_val or "test")
                        await self._session.wait(300)
                        self._write_result(ws, task["_row"], "PASS", "", (time.time() - t0) * 1000)
                        passed += 1
                    else:
                        self._write_result(ws, task["_row"], "FAIL", f"Unknown action: {action_name}", 0)
                        failed += 1
                except Exception as e:
                    self._write_result(ws, task["_row"], "FAIL", str(e)[:200], (time.time() - t0) * 1000)
                    failed += 1

                done += 1
                self._last_progress["done"] = done
                self._last_progress["passed"] = passed
                self._last_progress["failed"] = failed
                self._notify("step_done", {"done": done, "total": total_steps, "passed": passed, "failed": failed})

        # Save
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        out_path = self._xlsx_path.replace(".xlsx", f"_result_{ts}.xlsx")
        wb.save(out_path)

        # Stop browser
        video_path = ""
        try:
            video_path = await self._session.stop()
        except Exception:
            pass

        self._last_progress["result_path"] = out_path
        self._last_progress["video_path"] = video_path
        self._last_progress["done"] = done
        self._last_progress["passed"] = passed
        self._last_progress["failed"] = failed
        self._last_progress["total"] = total_steps
        self._last_progress["running"] = False
        self._last_progress["status"] = "finished"
        self._notify("finished", {
            "passed": passed, "failed": failed, "total": done,
            "result_path": out_path, "video_path": video_path,
        })
        return out_path

    @staticmethod
    def _strip_fp_suffix(url: str) -> str:
        """Remove #fp_* or #card_* or #search_* suffix, return base URL."""
        for marker in ("#fp_", "#card_", "#search_"):
            idx = url.find(marker)
            if idx >= 0:
                return url[:idx]
        return url

    @staticmethod
    def _is_fp_url(url: str) -> bool:
        return "#fp_" in url or "#card_" in url or "#search_" in url

    async def _replay_step(self, task: Dict[str, Any]):
        """Replay a step without recording results (to reach FP page state)."""
        action_name = str(task.get("action", "") or "")
        role = str(task.get("element_role", "") or "")
        input_val = str(task.get("input_value", "") or "")
        el_idx = int(task.get("element_index", 0) or 0)

        if action_name in ("点击", "click_target"):
            if el_idx > 0:
                await self._session.click_index(el_idx)
            else:
                await self._session.click_target({"role": role, "text_contains": str(task.get("element_text", "") or "")[:20]})
            await self._session.wait(500)
        elif action_name in ("输入", "type_target"):
            if el_idx > 0:
                await self._session.type_index(el_idx, input_val or "test")
            else:
                await self._session.type_target({"role": role, "placeholder": str(task.get("element_text", "") or "").lower()}, input_val or "test")
            await self._session.wait(300)

    @staticmethod
    def _write_result(ws, row: int, result: str, error: str, duration_ms: float):
        fill = PASS_FILL if result == "PASS" else FAIL_FILL
        for col_val, col_idx in [(result, HEADER_MAP["result"]), (error, HEADER_MAP["error_message"]), (f"{duration_ms:.0f}", HEADER_MAP["duration_ms"])]:
            cell = ws.cell(row=row, column=col_idx, value=col_val)
            cell.fill = fill
            cell.font = BOLD_FONT

    def _notify(self, event: str, data: Dict[str, Any]):
        if self._progress_callback:
            try:
                self._progress_callback(event, data)
            except Exception:
                pass
