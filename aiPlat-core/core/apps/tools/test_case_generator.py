"""
TestCaseGenerator — 页面分析 → 测试用例 Excel

复用 BrowserTestEngine 的页面发现 + 元素分析能力，生成结构化 Excel。
"""
from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .browser import _BrowserSession
from .browser_test_engine import ActionGenerator, PageDiscoverer, TestConfig


@dataclass
class TestCaseRow:
    case_id: str = ""
    case_title: str = ""
    precondition: str = ""
    case_type: str = "正向"
    page_url: str = ""
    step_no: int = 0
    current_page: str = ""
    target_page: str = ""
    element_role: str = ""
    element_text: str = ""
    element_type: str = ""
    element_index: int = 0
    action: str = ""
    input_value: str = ""
    expected_behavior: str = ""
    status: str = "PENDING"
    result: str = ""
    error_message: str = ""
    duration_ms: str = ""


HEADERS = [
    "case_id", "case_title", "precondition", "case_type",
    "page_url", "step_no", "current_page", "target_page",
    "element_role", "element_text", "element_type", "element_index",
    "action", "input_value", "expected_behavior", "status",
    "result", "error_message", "duration_ms",
]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PENDING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")


class TestCaseGenerator:
    """页面遍历分析 → 生成测试用例 Excel。"""

    def __init__(self, config: TestConfig):
        self._config = config
        self._session: Optional[_BrowserSession] = None
        self._discoverer: Optional[PageDiscoverer] = None
        self._generator = ActionGenerator()
        self._rows: List[TestCaseRow] = []
        self._visited: Set[str] = set()
        self._step_counter = 0
        self._case_counter = 0
        self._stop_requested = False

    def stop(self):
        self._stop_requested = True

    def _next_case(self) -> str:
        self._case_counter += 1
        return f"TC{self._case_counter:03d}"

    async def generate(self, output_path: str = "") -> str:
        """递归遍历页面并生成 xlsx 文件。每个元素链路完整探索后再进入下个链路。"""
        self._session = _BrowserSession()
        await self._session._lazy_start(headless=True)
        self._discoverer = PageDiscoverer(self._session, self._config)

        # Login if configured
        if self._config.login_url and self._config.accounts:
            from .browser_test_engine import SessionManager
            sm = SessionManager(self._session, self._config)
            if not await sm.ensure_login():
                raise RuntimeError("Login failed with all accounts")

        routes = self._config.routes or self._default_routes()
        content_fp: Set[str] = set()  # fingerprint dedup

        for route in routes:
            if self._stop_requested:
                break
            base_url = self._config.base_url
            norm = self._discoverer.normalize_url(
                f"{base_url}{route}" if not route.startswith("http") else route
            )
            await self._explore_page(norm, 0, content_fp)

        await self._session.stop()
        return self._write_xlsx(output_path)

    async def _explore_page(self, norm: str, depth: int, content_fp: Set[str],
                              case_id: str = "", case_title: str = "",
                              precondition: str = "", case_type: str = "正向"):
        """递归探索一个页面: 逐元素记录+即时探索完整链路, 再继续下个元素."""
        if self._stop_requested or len(self._rows) >= 500 or depth > self._config.max_recursion_depth:
            return
        base_url = self._config.base_url

        if not case_id:
            case_id = self._next_case()
            case_title = self._infer_case_title(norm)
            precondition = self._infer_precondition(norm)

        try:
            await self._session.goto(norm)
            await self._session.wait(2000)
        except Exception:
            return

        els_data = await self._session.list_elements(100)
        elements = els_data.get("elements", [])
        if not elements:
            return

        from .browser_test_engine import BrowserTestEngine

        # ── Step 1: Record cards first, then explore their popups ──
        await self._record_and_explore_cards(norm, elements, depth, content_fp, case_id)

        # ── Step 2: For each standard element, record → click → explore chain → go back ──
        for el in elements:
            if self._stop_requested or len(self._rows) >= 500:
                break
            await self._record_elements(norm, [el], skip_links=False, case_id=case_id, case_title=case_title, precondition=precondition, case_type=case_type)

            role = el.get("role", "")
            if role not in ("button", "link"):
                continue
            if depth >= self._config.max_recursion_depth:
                continue

            # Each clickable element starts a new sub-case (separate test chain)
            el_text = (el.get("visible_text", "") or el.get("name", "") or "").strip()[:30]
            sub_case = self._next_case()
            sub_title = self._infer_click_case_title(role, el_text, norm)
            sub_precond = self._infer_precondition(norm)

            try:
                await self._session.click_index(el["index"])
                await self._session.wait(1500)

                cur_hash = await self._session.evaluate("window.location.hash")
                cur_hash = str(cur_hash or "").strip()
                old_hash = "#" + norm.split("#")[1] if "#" in norm else ""

                if cur_hash and cur_hash != old_hash:
                    nn = self._discoverer.normalize_url(f"{base_url}{cur_hash}")
                    if not self._should_skip_url(nn):
                        await self._explore_page(nn, depth + 1, content_fp, sub_case, sub_title, sub_precond)
                else:
                    new_els = (await self._session.list_elements(100)).get("elements", [])
                    if new_els:
                        fp = BrowserTestEngine._compute_fingerprint(new_els)
                        old_fp = BrowserTestEngine._compute_fingerprint(elements)
                        if fp != old_fp and fp not in content_fp:
                            content_fp.add(fp)
                            fp_url = f"{norm}#fp_{len(content_fp)}"
                            await self._record_elements(fp_url, new_els, skip_links=True, case_id=sub_case, case_title=sub_title, precondition=sub_precond)
                            await self._explore_form_steps(fp_url, new_els, depth + 1, content_fp, sub_case, sub_title, sub_precond)

                await self._session.goto(norm)
                await self._session.wait(1000)
            except Exception:
                pass

        # ── Step 3: Search discovery (last) ──
        await self._discover_search(norm, elements, content_fp, case_id)

    async def _record_and_explore_cards(self, norm: str, elements: List[Dict[str, Any]],
                                         depth: int, content_fp: Set[str], case_id: str = ""):
        """逐卡片记录到主页, 并点击探索弹窗链路."""
        try:
            cards_data = await self._session.evaluate("""() => {
                const r = [];
                document.querySelectorAll('[class*=cursor-pointer]').forEach((el, i) => {
                    const rect = el.getBoundingClientRect();
                    if (rect.width > 100 && rect.height > 50) {
                        const h4 = el.querySelector('h4,h3,h2');
                        r.push({
                            index: i + 1000,
                            text: (el.textContent || '').trim().slice(0, 60),
                            title: h4 ? (h4.textContent || '').trim().slice(0, 40) : ((el.textContent || '').trim().split('\\n')[0].slice(0, 40)),
                        });
                    }
                });
                return r;
            }""")
            if not cards_data:
                return

            from .browser_test_engine import BrowserTestEngine

            for cd in cards_data:
                if self._stop_requested or len(self._rows) >= 500:
                    break
                title = cd.get("title", cd.get("text", "card"))
                if not title:
                    continue
                card_case = self._next_case()
                card_title = f"职位卡片「{title[:40]}」测试"
                card_precond = self._infer_precondition(norm)
                # Record card row on main page
                self._rows.append(TestCaseRow(
                    case_id=card_case, case_title=card_title,
                    precondition=card_precond, case_type="正向",
                    page_url=norm, step_no=self._step_counter + 1,
                    current_page="募集職種ページ", target_page="职位详情弹窗",
                    element_role="card", element_text=title[:200],
                    element_type="clickable-div", element_index=cd["index"],
                    action="点击", input_value="",
                    expected_behavior=f"【募集職種ページ】点职位卡片「{title[:40]}」 → 展示详情弹窗",
                ))
                self._step_counter += 1

                # Click this specific card to explore popup chain
                try:
                    title_js = title.replace("'", "\\'")
                    result = await self._session.evaluate(f"""() => {{
                        const els = document.querySelectorAll('[class*="cursor-pointer"]');
                        for (const el of els) {{
                            const r = el.getBoundingClientRect();
                            if (r.width > 100 && r.height > 50 && (el.textContent || '').includes('{title_js}')) {{
                                el.scrollIntoView({{behavior:'instant'}});
                                el.click();
                                return 'clicked';
                            }}
                        }}
                        return 'none';
                    }}""")
                    if result == 'clicked':
                        await self._session.wait(1000)
                        new_els = (await self._session.list_elements(50)).get("elements", [])
                        if new_els:
                            fp = BrowserTestEngine._compute_fingerprint(new_els)
                            old_fp = BrowserTestEngine._compute_fingerprint(elements)
                            if fp != old_fp and fp not in content_fp:
                                content_fp.add(fp)
                                card_url = f"{norm}#card_{len(content_fp)}"
                                await self._record_elements(card_url, new_els, skip_links=True, case_id=card_case, case_title=card_title, precondition=card_precond)
                        await self._session.goto(norm)
                        await self._session.wait(800)
                except Exception:
                    pass
        except Exception:
            pass

    async def _explore_form_steps(self, fp_url: str, elements: List[Dict[str, Any]],
                                   depth: int, content_fp: Set[str],
                                   case_id: str = "", case_title: str = "",
                                   precondition: str = "", case_type: str = "正向"):
        """递归探索表单的多步流转."""
        for _ in range(8):
            if self._stop_requested:
                break
            nav_btns = [e for e in elements
                        if e.get("role") == "button"
                        and any(kw in (e.get("visible_text", "") or "").lower()
                               for kw in ("submit", "次へ", "next", "送信"))]
            if not nav_btns:
                break
            try:
                btn = nav_btns[0]
                await self._session.click_index(btn["index"])
                await self._session.wait(1000)

                next_els = (await self._session.list_elements(100)).get("elements", [])
                if not next_els:
                    break
                from .browser_test_engine import BrowserTestEngine
                next_fp = BrowserTestEngine._compute_fingerprint(next_els)
                cur_fp = BrowserTestEngine._compute_fingerprint(elements)
                if next_fp != cur_fp and next_fp not in content_fp:
                    content_fp.add(next_fp)
                    fp_url = f"{fp_url.split('#fp_')[0]}#fp_{len(content_fp)}"
                    await self._record_elements(fp_url, next_els, skip_links=True, case_id=case_id, case_title=case_title, precondition=precondition, case_type=case_type)
                    elements = next_els
                else:
                    break
            except Exception:
                break
            nav_btns = [e for e in elements
                        if e.get("role") == "button"
                        and any(kw in (e.get("visible_text", "") or "").lower()
                               for kw in ("submit", "次へ", "next", "送信"))]
            if not nav_btns:
                break
            try:
                btn = nav_btns[0]
                await self._session.click_index(btn["index"])
                await self._session.wait(1000)

                next_els = (await self._session.list_elements(100)).get("elements", [])
                if not next_els:
                    break
                from .browser_test_engine import BrowserTestEngine
                next_fp = BrowserTestEngine._compute_fingerprint(next_els)
                cur_fp = BrowserTestEngine._compute_fingerprint(elements)
                if next_fp != cur_fp and next_fp not in content_fp:
                    content_fp.add(next_fp)
                    fp_url = f"{fp_url.split('#fp_')[0]}#fp_{len(content_fp)}"
                    await self._record_elements(fp_url, next_els, skip_links=True)
                    elements = next_els  # continue to next step
                else:
                    break
            except Exception:
                break

    async def _discover_search(self, norm: str, elements: List[Dict[str, Any]],
                                content_fp: Set[str], case_id: str = ""):
        """Type into search_input and hit Enter to discover filtered results."""
        if self._stop_requested:
            return
        try:
            await self._session.goto(norm)
            await self._session.wait(1000)
            await self._session.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await self._session.wait(500)

            current_els = (await self._session.list_elements(50)).get("elements", [])
            search_els = [e for e in current_els if e.get("role") == "search_input"]
            if not search_els:
                return
            se = search_els[0]

            old_texts = await self._get_visible_card_texts()
            await self._session.type_index(se["index"], "IT")
            await self._session.wait(500)
            await self._session.evaluate("""() => {
                const el = document.activeElement;
                if (el) {
                    el.dispatchEvent(new KeyboardEvent('keydown', {key:'Enter', code:'Enter', keyCode:13, bubbles:true}));
                    el.dispatchEvent(new KeyboardEvent('keyup', {key:'Enter', code:'Enter', keyCode:13, bubbles:true}));
                }
            }""")
            await self._session.wait(2000)

            try:
                new_texts = await self._get_visible_card_texts()
            except Exception:
                new_texts = ""
            if new_texts != old_texts and new_texts not in content_fp:
                content_fp.add(new_texts)
                fp_url = f"{norm}#search_{len(content_fp)}"
                new_els = (await self._session.list_elements(50)).get("elements", [])
                if new_els:
                    srch_case = self._next_case()
                    srch_title = "搜索过滤功能测试"
                    srch_precond = self._infer_precondition(norm)
                    await self._record_elements(fp_url, new_els, skip_links=False, case_id=srch_case, case_title=srch_title, precondition=srch_precond)

            await self._session.goto(norm)
            await self._session.wait(1000)
        except Exception:
            pass

    async def _get_visible_card_texts(self) -> str:
        """Return hash of cursor-pointer card text content (all, not just viewport)."""
        try:
            txt = await self._session.evaluate("""() => {
                const texts = [];
                document.querySelectorAll('[class*=cursor-pointer]').forEach(el => {
                    const r = el.getBoundingClientRect();
                    if (r.width > 100 && r.height > 50) {
                        texts.push((el.textContent || '').trim().slice(0, 120));
                    }
                });
                return texts.sort().join('|');
            }""")
            return txt or ""
        except Exception:
            return ""

    async def _record_elements(self, url: str, elements: List[Dict[str, Any]], skip_links: bool = False,
                                case_id: str = "", case_title: str = "",
                                precondition: str = "", case_type: str = "正向"):
        """Generate TestCaseRow for each element on a page."""

        for el in elements:
            role = el.get("role", "")
            if skip_links and role == "link":
                continue
            tp = el.get("type", "")
            text = (el.get("visible_text") or el.get("name") or el.get("placeholder") or f"[{tp}]" or "").strip()
            idx = el.get("index", 0)
            self._step_counter += 1

            action_info = self._generator._role_to_action(role, el.get("tag", ""), el)
            if not action_info:
                continue

            act = action_info.get("action", "")
            action_name = "点击" if act == "click_target" else "输入" if act == "type_target" else act
            input_val = action_info.get("text", "") if act == "type_target" else ""
            if act == "type_target" and not input_val:
                input_val = self._generator._smart_text(el)

            cur_page = self._page_name(url)
            target = self._target_page(act, role, text, url)
            behavior = self._infer_behavior(act, role, text, input_val, url)

            self._rows.append(TestCaseRow(
                case_id=case_id,
                case_title=case_title,
                precondition=precondition,
                case_type=case_type,
                page_url=url,
                step_no=self._step_counter,
                current_page=cur_page,
                target_page=target,
                element_role=role,
                element_text=text[:200],
                element_type=tp,
                element_index=idx,
                action=action_name,
                input_value=input_val,
                expected_behavior=behavior,
            ))

    @staticmethod
    def _infer_case_title(url: str) -> str:
        page = TestCaseGenerator._page_name(url)
        if page == "募集職種ページ":
            return "募集職種ページ基本操作と要素探索"
        return f"{page}の要素テスト"

    @staticmethod
    def _infer_click_case_title(role: str, text: str, url: str) -> str:
        page = TestCaseGenerator._page_name(url)
        if role == "button":
            return f"【{page}】ボタン「{text[:30]}」テスト"
        if role == "link":
            return f"【{page}】リンク「{text[:30]}」からの遷移テスト"
        return f"【{page}】{text[:30]} 操作テスト"

    @staticmethod
    def _infer_precondition(url: str) -> str:
        page = TestCaseGenerator._page_name(url)
        return f"{page}（{url[:60]}）が表示されていること"

    def _should_skip_url(self, url: str) -> bool:
        import re
        for p in self._config.exclude_patterns:
            if re.search(p, url):
                return True
        if self._config.include_patterns:
            if not any(re.search(p, url) for p in self._config.include_patterns):
                return True
        return False

    def _write_xlsx(self, output_path: str = "") -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Header
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # Data rows
        for r, row in enumerate(self._rows, 2):
            values = [
                row.case_id, row.case_title, row.precondition, row.case_type,
                row.page_url, row.step_no, row.current_page, row.target_page,
                row.element_role, row.element_text, row.element_type, row.element_index,
                row.action, row.input_value, row.expected_behavior,
                row.status, row.result, row.error_message, row.duration_ms,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if col == HEADERS.index("status") + 1:
                    cell.fill = PENDING_FILL

        # Column widths
        widths = [10, 35, 40, 8, 45, 8, 20, 20, 14, 30, 12, 8, 8, 25, 40, 12, 10, 35, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze top row
        ws.freeze_panes = "A2"

        if not output_path:
            ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                tempfile.gettempdir(), f"browser_test_cases_{ts}.xlsx"
            )
        wb.save(output_path)
        return output_path

    @staticmethod
    def _page_name(url: str) -> str:
        if "card_" in url:
            return "职位详情弹窗"
        if "search_" in url:
            return "搜索过滤结果"
        if "fp_2" in url:
            return "応募表单-第1步"
        if "fp_3" in url:
            return "応募表单-第2步"
        if "fp_4" in url:
            return "応募表单-第3步(完整填表)"
        if "fp_5" in url:
            return "応募表单-第4步(退回)"
        if "fp_" in url:
            n = url.split("fp_")[1].split("#")[0].split("_")[0] if "#" in url else ""
            return f"応募表单-步骤{n}"
        if "#/careers" in url or "/careers" in url.split("#")[-1] if "#" in url else "":
            return "募集職種ページ"
        if "login" in url:
            return "登录页"
        if "register" in url:
            return "注册页"
        return "页面"

    @staticmethod
    def _target_page(act: str, role: str, text: str, page_url: str) -> str:
        """Predict the target page after this action."""
        if act != "click_target":
            return ""
        if role == "link":
            return "目标页面"
        if "search_input" in role or "検索" in text:
            return "搜索过滤结果"
        if role == "card":
            return "职位详情弹窗"
        # Text-based inference
        t = (text or "").lower()
        if "駐在企業" in t or "驻在企業" in t:
            return "駐在企業一覧"
        if "今すぐ応募" in t or "応募" in t:
            return "応募表单"
        if "職種一覧" in t or "职種一覧" in t:
            return "职位列表"
        if "戻る" in t or "返回" in t:
            return "返回上一页"
        if "次へ" in t or "next" in t:
            return "表单下一步"
        if "submit" in t or "送信" in t or "登録申請" in t:
            return "提交表单/确认页"
        if "ログイン" in t or "login" in t:
            return "登录后页面"
        if "面接" in t:
            return "面接入室画面"
        if "企業" in t or "company" in t:
            return "企業関連ページ"
        if "利用規約" in t:
            return "利用規約画面"
        if "プライバシー" in t or "privacy" in t:
            return "プライバシーポリシー画面"
        if "パスワード" in t or "password" in t or "忘れ" in t:
            return "パスワード関連画面"
        if "新規" in t or "登録" in t or "申請" in t:
            return "登録/申請画面"
        return "次画面/弹窗"

    @staticmethod
    def _infer_behavior(act: str, role: str, text: str, input_val: str, page_url: str = "") -> str:
        page = TestCaseGenerator._page_name(page_url)
        if act == "click_target":
            if role == "link":
                return f"【{page}】点链接「{text[:30]}」 → 导航到新页面"
            if role == "button":
                return f"【{page}】点「{text[:30]}」 → 触发表单/弹窗"
            if role == "card":
                return f"【{page}】点职位卡片「{text[:30]}」 → 展示详情弹窗"
            if role == "text_input":
                return f"【{page}】选择「{text[:15]}」"
            return f"【{page}】点击 {role}「{text[:30]}」"
        if act == "type_target":
            return f"【{page}】输入「{input_val}」"
        return f"【{page}】{act}"

    @staticmethod
    def _default_routes() -> List[str]:
        return ["/"]

    @staticmethod
    def normalize_url(url: str) -> str:
        base = url.split("?")[0]
        if "#" in base:
            parts = base.split("#", 1)
            return f"{parts[0].rstrip('/')}#{parts[1]}"
        return base.rstrip("/")
